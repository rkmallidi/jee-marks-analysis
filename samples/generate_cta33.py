"""
Generate questions Excel for CTA-33 (JEE-ADV-2021-P2) from parsed key sheet.

JEE Advanced 2021 Paper 2 pattern (19 questions per subject, 57 total):
  Section 1  Q1-6   per sub : MCQ_MULTI       +4  -2   (partial credit)
  Section 2  Q7-12  per sub : NUMERICAL_DECIMAL +2   0
  Section 3  Q13-16 per sub : SCQ (paragraph)  +3  -1
  Section 4  Q17-19 per sub : NUMERICAL_DECIMAL +4   0

Marks per subject: 6×4 + 6×2 + 4×3 + 3×4 = 24+12+12+12 = 60  →  Total = 180
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COLS = ["qno","subject","topic","sub_topic","difficulty",
        "question_type","marks","negative_marks","correct_answer","partial_marks","is_deleted"]

rows = [
    # ── PHYSICS  (Q1-19) ───────────────────────────────────────────────────────
    # Section 1: MCQ_MULTI +4 -2
    [1,  "Physics","Electromagnetic Induction","LR Circuit - Current Growth",          "","MCQ_MULTI",     4,2,"AD",  0,"N"],
    [2,  "Physics","Electromagnetic Induction","Circuit Symmetry",                     "","MCQ_MULTI",     4,2,"BCD", 0,"N"],
    [3,  "Physics","Electromagnetic Induction","Charged Particle in Magnetic Field",   "","MCQ_MULTI",     4,2,"BC",  0,"N"],
    [4,  "Physics","Electromagnetic Induction","Torque on Current Loop",               "","MCQ_MULTI",     4,2,"AD",  0,"N"],
    [5,  "Physics","Electromagnetic Induction","Charged Disc in Changing B Field",     "","MCQ_MULTI",     4,2,"ACD", 0,"N"],
    [6,  "Physics","Electromagnetic Induction","Motional EMF - Rotating Rod",          "","MCQ_MULTI",     4,2,"AB",  0,"N"],
    # Section 2: NUMERICAL_DECIMAL +2 0
    [7,  "Physics","Waves","Tension in Vibrating String - T1",                        "","NUMERICAL_DECIMAL",2,0,"6",  0,"N"],
    [8,  "Physics","Waves","Tension in Vibrating String - T2",                        "","NUMERICAL_DECIMAL",2,0,"14", 0,"N"],
    [9,  "Physics","Electromagnetic Induction","Eddy Current Power Dissipation",       "","NUMERICAL_DECIMAL",2,0,"64", 0,"N"],
    [10, "Physics","Electromagnetic Induction","Power Scaling with Dimension",         "","NUMERICAL_DECIMAL",2,0,"16", 0,"N"],
    [11, "Physics","Electromagnetic Induction","Induced Angular Acceleration",         "","NUMERICAL_DECIMAL",2,0,"2",  0,"N"],
    [12, "Physics","Electromagnetic Induction","Induced Velocity of Charged Disc",     "","NUMERICAL_DECIMAL",2,0,"0.5",0,"N"],
    # Section 3: SCQ (paragraph) +3 -1
    [13, "Physics","Electromagnetic Induction","Force on Triangular Current Loop",     "","SCQ",           3,1,"B",  0,"N"],
    [14, "Physics","Electromagnetic Induction","SHM of Current Loop in B Field",       "","SCQ",           3,1,"C",  0,"N"],
    [15, "Physics","Electromagnetic Induction","Rotating Charge - Torque",             "","SCQ",           3,1,"C",  0,"N"],
    [16, "Physics","Electromagnetic Induction","Rotating Charge - Angular Momentum",   "","SCQ",           3,1,"B",  0,"N"],
    # Section 4: NUMERICAL_DECIMAL +4 0
    [17, "Physics","Electromagnetic Induction","Net Magnetic Flux",                    "","NUMERICAL_DECIMAL",4,0,"3",  0,"N"],
    [18, "Physics","Electromagnetic Induction","Magnetic Field of Rotating Charge",    "","NUMERICAL_DECIMAL",4,0,"4",  0,"N"],
    [19, "Physics","Electromagnetic Induction","Ampere's Law Path Integral",           "","NUMERICAL_DECIMAL",4,0,"4",  0,"N"],

    # ── CHEMISTRY  (Q20-38) ────────────────────────────────────────────────────
    # Section 1: MCQ_MULTI +4 -2
    [20, "Chemistry","Metallurgy","Hydrogen as Reducing Agent",                        "","MCQ_MULTI",     4,2,"ABD", 0,"N"],
    [21, "Chemistry","Metallurgy","General Principles of Metallurgy",                  "","MCQ_MULTI",     4,2,"ABCD",0,"N"],
    [22, "Chemistry","Metallurgy","Copper Extraction Reactions",                       "","MCQ_MULTI",     4,2,"BCD", 0,"N"],
    [23, "Chemistry","Ionic Equilibrium","Solubility of AgCl in KBr",                 "","MCQ_MULTI",     4,2,"CD",  0,"N"],
    [24, "Chemistry","Ionic Equilibrium","Selective Precipitation of AgBr",            "","MCQ_MULTI",     4,2,"ACD", 0,"N"],
    [25, "Chemistry","Ionic Equilibrium","Acid-Base Indicators",                       "","MCQ_MULTI",     4,2,"AD",  0,"N"],
    # Section 2: NUMERICAL_DECIMAL +2 0
    [26, "Chemistry","Metallurgy","Cu2S Oxidation - Moles of O2",                     "","NUMERICAL_DECIMAL",2,0,"960",0,"N"],
    [27, "Chemistry","Metallurgy","Reaction Products Identification",                  "","NUMERICAL_DECIMAL",2,0,"10", 0,"N"],
    [28, "Chemistry","Metallurgy","Mass Calculation - Hectogram",                      "","NUMERICAL_DECIMAL",2,0,"7.44",0,"N"],
    [29, "Chemistry","Metallurgy","Cyanide Complex Stoichiometry",                     "","NUMERICAL_DECIMAL",2,0,"8",  0,"N"],
    [30, "Chemistry","Ionic Equilibrium","Weak Acid Equilibrium - H+ Concentration",  "","NUMERICAL_DECIMAL",2,0,"5",  0,"N"],
    [31, "Chemistry","Ionic Equilibrium","pH Change with Dilution",                    "","NUMERICAL_DECIMAL",2,0,"6",  0,"N"],
    # Section 3: SCQ (paragraph) +3 -1
    [32, "Chemistry","Metallurgy","Zinc Blende Extraction Process",                    "","SCQ",           3,1,"D",  0,"N"],
    [33, "Chemistry","Metallurgy","Ellingham Diagram - Temperature Effect",            "","SCQ",           3,1,"C",  0,"N"],
    [34, "Chemistry","Ionic Equilibrium","Charge Balance in HCl Solution",             "","SCQ",           3,1,"B",  0,"N"],
    [35, "Chemistry","Ionic Equilibrium","Mixture of HCl and Weak Acid",              "","SCQ",           3,1,"A",  0,"N"],
    # Section 4: NUMERICAL_DECIMAL +4 0
    [36, "Chemistry","Metallurgy","Mg Displacement - Feasibility",                     "","NUMERICAL_DECIMAL",4,0,"0",  0,"N"],
    [37, "Chemistry","Metallurgy","Moles of Cu from Electrolysis",                     "","NUMERICAL_DECIMAL",4,0,"99", 0,"N"],
    [38, "Chemistry","Ionic Equilibrium","Ksp of Pb(OH)2 - Solubility Ratio",         "","NUMERICAL_DECIMAL",4,0,"4",  0,"N"],

    # ── MATHEMATICS  (Q39-57) ──────────────────────────────────────────────────
    # Section 1: MCQ_MULTI +4 -2
    [39, "Mathematics","Functions","Greatest Integer Function - Composition",          "","MCQ_MULTI",     4,2,"AC",  0,"N"],
    [40, "Mathematics","Definite Integrals","King's Rule - Integral Properties",       "","MCQ_MULTI",     4,2,"ACD", 0,"N"],
    [41, "Mathematics","Definite Integrals","Inverse Function Integration",            "","MCQ_MULTI",     4,2,"BD",  0,"N"],
    [42, "Mathematics","Polynomials","Polynomial Division and Remainder",              "","MCQ_MULTI",     4,2,"ABD", 0,"N"],
    [43, "Mathematics","Definite Integrals","Substitution Method",                     "","MCQ_MULTI",     4,2,"ABCD",0,"N"],
    [44, "Mathematics","Definite Integrals","Integral Ratio - Reduction Formula",      "","MCQ_MULTI",     4,2,"AC",  0,"N"],
    # Section 2: NUMERICAL_DECIMAL +2 0
    [45, "Mathematics","Area Under Curves","Area Between Quadratic Curves - R(0)",     "","NUMERICAL_DECIMAL",2,0,"116",0,"N"],
    [46, "Mathematics","Area Under Curves","Area Between Quadratic Curves - Ellipse",  "","NUMERICAL_DECIMAL",2,0,"128",0,"N"],
    [47, "Mathematics","Area Under Curves","Triangle Area with Geometric Constraint",  "","NUMERICAL_DECIMAL",2,0,"5",  0,"N"],
    [48, "Mathematics","Area Under Curves","Geometric Region - Parameter a",           "","NUMERICAL_DECIMAL",2,0,"10", 0,"N"],
    [49, "Mathematics","Area Under Curves","Area Between Ellipse and Line - A1",       "","NUMERICAL_DECIMAL",2,0,"10", 0,"N"],
    [50, "Mathematics","Area Under Curves","Area Between Ellipse and Line - A2",       "","NUMERICAL_DECIMAL",2,0,"20", 0,"N"],
    # Section 3: SCQ (paragraph) +3 -1
    [51, "Mathematics","Definite Integrals","Partial Fractions - p+q+r",               "","SCQ",           3,1,"D",  0,"N"],
    [52, "Mathematics","Definite Integrals","Partial Fractions - Floor of Sum",        "","SCQ",           3,1,"A",  0,"N"],
    [53, "Mathematics","Definite Integrals","Integration by Parts - g(x)",             "","SCQ",           3,1,"B",  0,"N"],
    [54, "Mathematics","Definite Integrals","Integration by Parts - Result",           "","SCQ",           3,1,"B",  0,"N"],
    # Section 4: NUMERICAL_DECIMAL +4 0
    [55, "Mathematics","Definite Integrals","Floor Function Integral - 10's Digit",    "","NUMERICAL_DECIMAL",4,0,"8",  0,"N"],
    [56, "Mathematics","Definite Integrals","Inverse Tangent Integral",                "","NUMERICAL_DECIMAL",4,0,"2",  0,"N"],
    [57, "Mathematics","Definite Integrals","King's Rule - a+b Value",                 "","NUMERICAL_DECIMAL",4,0,"5",  0,"N"],
]

# ── Build workbook ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Questions"

hdr_fill = PatternFill("solid", fgColor="1E3A5F")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
thin      = Side(style="thin", color="CCCCCC")
border    = Border(left=thin, right=thin, top=thin, bottom=thin)

for ci, col in enumerate(COLS, 1):
    cell = ws.cell(row=1, column=ci, value=col)
    cell.font      = hdr_font
    cell.fill      = hdr_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = border

SUBJ_FILL = {
    "Physics":     PatternFill("solid", fgColor="EAFAF1"),
    "Chemistry":   PatternFill("solid", fgColor="FEF9E7"),
    "Mathematics": PatternFill("solid", fgColor="EBF5FB"),
}

for ri, row in enumerate(rows, 2):
    subj = row[1]
    fill = SUBJ_FILL.get(subj)
    for ci, val in enumerate(row, 1):
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.border    = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if fill:
            cell.fill = fill
        if ci in (2, 3, 4, 5, 7):
            cell.alignment = Alignment(horizontal="left", vertical="center")

widths = [6, 14, 26, 40, 12, 20, 8, 17, 16, 14, 11]
for ci, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

ws.row_dimensions[1].height = 22
ws.freeze_panes = "A2"

out = "samples/CTA-33_JEE-ADV-2021-P2_Questions.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Total rows : {len(rows)}")

total = sum(r[6] for r in rows)
print(f"Total max marks : {total}")
for subj in ("Physics","Chemistry","Mathematics"):
    s = sum(r[6] for r in rows if r[1]==subj)
    print(f"  {subj}: {s}")
