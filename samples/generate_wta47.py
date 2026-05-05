"""
Generate questions Excel for WTA-47 (JEE-ADV-2025-P1) from parsed key sheet.

Pattern (per subject, 16 questions):
  Section 1  Q1-4   : SCQ          +3  -1   partial=0
  Section 2  Q5-7   : MCQ_MULTI    +4  -2   partial=0  (JEE Adv partial scoring)
  Section 3  Q8-13  : NUMERICAL_DECIMAL +4  0
  Section 4  Q14-16 : SCQ          +4  -1   partial=0
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COLS = ["qno","subject","topic","sub_topic","difficulty",
        "question_type","marks","negative_marks","correct_answer","partial_marks","is_deleted"]

rows = [
    # ── MATHEMATICS ────────────────────────────────────────────────────────────
    # Section 1: SCQ +3 -1
    [1,  "Mathematics","Differential Equations","First Order Linear ODE",     "","SCQ",           3,1,"D",  0,"N"],
    [2,  "Mathematics","Differential Equations","Separable ODE",               "","SCQ",           3,1,"C",  0,"N"],
    [3,  "Mathematics","Differential Equations","First Order Linear ODE",     "","SCQ",           3,1,"A",  0,"N"],
    [4,  "Mathematics","Differential Equations","Integral Equations",          "","SCQ",           3,1,"A",  0,"N"],
    # Section 2: MCQ_MULTI +4 -2
    [5,  "Mathematics","Differential Equations","Applications of ODE",         "","MCQ_MULTI",     4,2,"BCD",0,"N"],
    [6,  "Mathematics","Differential Equations","Tangent and Normal",          "","MCQ_MULTI",     4,2,"AC", 0,"N"],
    [7,  "Mathematics","Differential Equations","Homogeneous ODE",             "","MCQ_MULTI",     4,2,"AB", 0,"N"],
    # Section 3: NUMERICAL_DECIMAL +4 0
    [8,  "Mathematics","Differential Equations","Bernoulli's Equation",        "","NUMERICAL_DECIMAL",4,0,"4.8",0,"N"],
    [9,  "Mathematics","Differential Equations","Integral Equations",          "","NUMERICAL_DECIMAL",4,0,"9",  0,"N"],
    [10, "Mathematics","Differential Equations","Higher Order ODE",            "","NUMERICAL_DECIMAL",4,0,"1",  0,"N"],
    [11, "Mathematics","Differential Equations","Exact ODE",                   "","NUMERICAL_DECIMAL",4,0,"1",  0,"N"],
    [12, "Mathematics","Differential Equations","First Order Linear ODE",     "","NUMERICAL_DECIMAL",4,0,"1",  0,"N"],
    [13, "Mathematics","Differential Equations","Population Models",           "","NUMERICAL_DECIMAL",4,0,"3",  0,"N"],
    # Section 4: SCQ +4 -1
    [14, "Mathematics","Differential Equations","Order and Degree",            "","SCQ",           4,1,"D",  0,"N"],
    [15, "Mathematics","Differential Equations","Order and Degree",            "","SCQ",           4,1,"C",  0,"N"],
    [16, "Mathematics","Differential Equations","Integral Equations",          "","SCQ",           4,1,"B",  0,"N"],

    # ── PHYSICS ────────────────────────────────────────────────────────────────
    # Section 1: SCQ +3 -1
    [17, "Physics","Electromagnetic Induction","LR Circuits - Energy",        "","SCQ",           3,1,"C",  0,"N"],
    [18, "Physics","Electromagnetic Induction","Moving Rod - Energy Conservation","","SCQ",        3,1,"B",  0,"N"],
    [19, "Physics","Electromagnetic Induction","Rotating Magnetic Dipole",    "","SCQ",           3,1,"C",  0,"N"],
    [20, "Physics","Electromagnetic Induction","Thevenin Equivalent LR",      "","SCQ",           3,1,"B",  0,"N"],
    # Section 2: MCQ_MULTI +4 -2
    [21, "Physics","Electromagnetic Induction","Kirchhoff's Laws with Inductors","","MCQ_MULTI",  4,2,"ACD",0,"N"],
    [22, "Physics","Electromagnetic Induction","Kirchhoff's Laws with Inductors","","MCQ_MULTI",  4,2,"ACD",0,"N"],
    [23, "Physics","Electromagnetic Induction","Moving Rod SHM",              "","MCQ_MULTI",     4,2,"BCD",0,"N"],
    # Section 3: NUMERICAL_DECIMAL +4 0
    [24, "Physics","Electromagnetic Induction","Torque on Rotating Disc",     "","NUMERICAL_DECIMAL",4,0,"2.5",0,"N"],
    [25, "Physics","Electromagnetic Induction","Flux Conservation",           "","NUMERICAL_DECIMAL",4,0,"5",  0,"N"],
    [26, "Physics","Electromagnetic Induction","Magnetic Energy Density",     "","NUMERICAL_DECIMAL",4,0,"7",  0,"N"],
    [27, "Physics","Electromagnetic Induction","Motional EMF",                "","NUMERICAL_DECIMAL",4,0,"9",  0,"N"],
    [28, "Physics","Electromagnetic Induction","Charge Flow",                 "","NUMERICAL_DECIMAL",4,0,"2",  0,"N"],
    [29, "Physics","Electromagnetic Induction","Electric Field - Changing Flux","","NUMERICAL_DECIMAL",4,0,"1",0,"N"],
    # Section 4: SCQ +4 -1
    [30, "Physics","Electromagnetic Induction","LR Circuit - Qualitative",    "","SCQ",           4,1,"C",  0,"N"],
    [31, "Physics","Electromagnetic Induction","Power in Circuits",           "","SCQ",           4,1,"C",  0,"N"],
    [32, "Physics","Electromagnetic Induction","Inductors - Conceptual",      "","SCQ",           4,1,"B",  0,"N"],

    # ── CHEMISTRY ──────────────────────────────────────────────────────────────
    # Section 1: SCQ +3 -1
    [33, "Chemistry","Metallurgy","Refining Methods",                         "","SCQ",           3,1,"C",  0,"N"],
    [34, "Chemistry","Metallurgy","Ellingham Diagram",                        "","SCQ",           3,1,"C",  0,"N"],
    [35, "Chemistry","Metallurgy","Calcination and Roasting",                 "","SCQ",           3,1,"B",  0,"N"],
    [36, "Chemistry","Metallurgy","Flux and Slag",                            "","SCQ",           3,1,"C",  0,"N"],
    # Section 2: MCQ_MULTI +4 -2
    [37, "Chemistry","Metallurgy","Extraction of Copper",                     "","MCQ_MULTI",     4,2,"ABD",0,"N"],
    [38, "Chemistry","Metallurgy","Copper Matte and Byproducts",              "","MCQ_MULTI",     4,2,"BC", 0,"N"],
    [39, "Chemistry","Metallurgy","Gases in Metallurgy",                      "","MCQ_MULTI",     4,2,"BC", 0,"N"],
    # Section 3: NUMERICAL_DECIMAL +4 0
    [40, "Chemistry","Metallurgy","Hydrometallurgy",                          "","NUMERICAL_DECIMAL",4,0,"4",  0,"N"],
    [41, "Chemistry","Metallurgy","Impurities in Metals",                     "","NUMERICAL_DECIMAL",4,0,"6",  0,"N"],
    [42, "Chemistry","Metallurgy","Concentration Calculations",               "","NUMERICAL_DECIMAL",4,0,"20", 0,"N"],
    [43, "Chemistry","Metallurgy","Percentage Calculations",                  "","NUMERICAL_DECIMAL",4,0,"50", 0,"N"],
    [44, "Chemistry","Metallurgy","Complex Compounds in Metallurgy",          "","NUMERICAL_DECIMAL",4,0,"5",  0,"N"],
    [45, "Chemistry","Metallurgy","Cyanide Process",                          "","NUMERICAL_DECIMAL",4,0,"6",  0,"N"],
    # Section 4: SCQ +4 -1
    [46, "Chemistry","Metallurgy","Hydrometallurgy Applications",             "","SCQ",           4,1,"D",  0,"N"],
    [47, "Chemistry","Metallurgy","Gases and Odor in Metallurgy",             "","SCQ",           4,1,"C",  0,"N"],
    [48, "Chemistry","Metallurgy","Flame Colors and Properties",              "","SCQ",           4,1,"A",  0,"N"],
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Questions"

# ── Header styling ────────────────────────────────────────────────────────────
hdr_fill = PatternFill("solid", fgColor="1E3A5F")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for ci, col in enumerate(COLS, 1):
    cell = ws.cell(row=1, column=ci, value=col)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# Subject fill colours
SUBJ_FILL = {
    "Mathematics": PatternFill("solid", fgColor="EBF5FB"),
    "Physics":     PatternFill("solid", fgColor="EAFAF1"),
    "Chemistry":   PatternFill("solid", fgColor="FEF9E7"),
}

# ── Data rows ─────────────────────────────────────────────────────────────────
for ri, row in enumerate(rows, 2):
    subj = row[1]
    fill = SUBJ_FILL.get(subj)
    for ci, val in enumerate(row, 1):
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if fill:
            cell.fill = fill
        # left-align text columns
        if ci in (2, 3, 4, 5, 7):
            cell.alignment = Alignment(horizontal="left", vertical="center")

# ── Column widths ─────────────────────────────────────────────────────────────
widths = [6, 14, 26, 36, 12, 20, 8, 17, 16, 14, 11]
for ci, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

ws.row_dimensions[1].height = 22
ws.freeze_panes = "A2"

out = "samples/WTA-47_JEE-ADV-2025-P1_Questions.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Total rows: {len(rows)}")

# Quick verification
total_marks = sum(r[6] for r in rows)
print(f"Total max marks: {total_marks}")
print(f"Math: {sum(r[6] for r in rows if r[1]=='Mathematics')}")
print(f"Phys: {sum(r[6] for r in rows if r[1]=='Physics')}")
print(f"Chem: {sum(r[6] for r in rows if r[1]=='Chemistry')}")
